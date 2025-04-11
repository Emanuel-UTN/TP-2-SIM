import os
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from libreria.random_numbers import (
    uniform_samples,
    poisson_samples,
    exponential_samples,
    normal_samples
)
from validador import (
    validar_ab_uniforme,
    validar_lambda,
    validar_stdev,
    validar_bins,
    validar_media
)

def visualizations(distribution_type: str, sample_size: int, bins: int = None) -> tuple:
    """
    Genera muestras aleatorias, crea un histograma, guarda los datos en Excel y devuelve los datos junto a kwargs
    útiles para funciones estadísticas posteriores.

    Parámetros:
    - distribution_type: Tipo de distribución ('u', 'p', 'e', 'n')
    - sample_size: Tamaño de muestra
    - bins: Cantidad de intervalos (opcional)

    Retorna:
    - data: Lista con los valores generados
    - kwargs: Diccionario con parámetros de la distribución para uso posterior
    """
    
    # Validar bins si no fue especificado
    if bins is None:
        bins = validar_bins()

    # Inicializar variables
    kwargs = {}
    data = []
    title = ""
    image_filename = f'temporal_{distribution_type}_distribution.png'

    # Generación de muestras según el tipo de distribución
    if distribution_type == 'u':  # Uniforme
        a, b = validar_ab_uniforme()
        data = uniform_samples(sample_size, a, b)
        title = f'Distribución Uniforme ({a}-{b})'

    elif distribution_type == 'p':  # Poisson
        mu = validar_lambda("poisson")
        data = poisson_samples(sample_size, mu)
        title = f'Distribución de Poisson (λ={mu})'
        bins = range(0, max(data) + 2)
        kwargs['mu'] = mu

    elif distribution_type == 'e':  # Exponencial
        lambd = validar_lambda("exponencial")
        data = exponential_samples(sample_size, lambd)
        title = f'Distribución Exponencial (λ={lambd})'
        kwargs['lambd'] = lambd

    elif distribution_type == 'n':  # Normal
        media = validar_media("normal")
        sigma = validar_stdev()
        data = normal_samples(sample_size, media, sigma)
        title = f'Distribución Normal (μ={media}, σ={sigma})'
        kwargs['media'] = media
        kwargs['sigma'] = sigma

    else:
        print("Tipo de distribución inválido. Elija 'u', 'p', 'e', o 'n'.")
        return [], {}

    # === Crear y guardar el histograma ===
    plt.figure(figsize=(8, 6))
    plt.hist(data, bins=bins, alpha=0.7)
    plt.title(title)
    plt.xlabel('Valor')
    plt.ylabel('Frecuencia')
    plt.tight_layout()
    plt.savefig(image_filename)
    plt.close()

    # === Guardar datos y gráfico en archivo Excel ===
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    excel_filename = os.path.join(output_dir, f'{distribution_type}_output.xlsx')

    # Guardar datos en hoja 'Datos'
    df = pd.DataFrame({f'{distribution_type}_valores': data})
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos')

    # Insertar gráfico en hoja nueva 'Gráfico'
    wb = load_workbook(excel_filename)
    ws_chart = wb.create_sheet('Gráfico')
    img = ExcelImage(image_filename)
    img.anchor = 'A1'
    ws_chart.add_image(img)
    wb.save(excel_filename)

    # Eliminar imagen temporal
    if os.path.exists(image_filename):
        os.remove(image_filename)

    # Mostrar ruta absoluta del archivo generado
    abs_path = os.path.abspath(excel_filename).replace(os.sep, '/').replace(' ', '%20')
    print(f"Archivo Excel creado en: file://{abs_path}")

    return data, kwargs